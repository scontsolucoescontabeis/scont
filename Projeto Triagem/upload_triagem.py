#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
upload_triagem.py
========================================
Lê a planilha triagem_atendimento.xlsx e faz o upload
dos dados para o Supabase (tabelas triagem_demandas,
triagem_resumo e triagem_conversas).

INSTALAÇÃO (apenas uma vez):
  pip install openpyxl requests

EXECUÇÃO MANUAL:
  python upload_triagem.py

AGENDAMENTO AUTOMÁTICO:
  Use o agendar_upload.bat ou o Agendador de Tarefas do Windows.
========================================
"""

import sys
import re
import json
import logging
import requests
from pathlib import Path
from datetime import datetime, timezone

# ── Configuração ──────────────────────────────────────────────────────────────

SUPABASE_URL = "https://dsdqwigopzrdmxtmhsez.supabase.co"
SUPABASE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRzZHF3aWdvcHpyZG14dG1oc2V6Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzY2ODM2OTUsImV4cCI6MjA5MjI1OTY5NX0"
    ".MPxbcKh6N_BNh0zTTb-jtNigQwCp-e6g3xboBbNbRmw"
)

# Caminho padrão da planilha — ajuste se necessário
XLSX_PATH = (
    Path.home()
    / "Desktop"
    / "captura_messenger"
    / "Triagem de Atendimentos - Gestta"
    / "triagem_atendimento.xlsx"
)

# ── Logging ───────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Supabase helpers ──────────────────────────────────────────────────────────

HEADERS = {
    "apikey":        SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type":  "application/json",
    "Prefer":        "return=minimal",
}


def _url(table: str) -> str:
    return f"{SUPABASE_URL}/rest/v1/{table}"


def sb_delete_all(table: str, filter_col: str = "id", min_val: int = 0) -> None:
    """Apaga todas as linhas de uma tabela via filtro gte."""
    r = requests.delete(
        _url(table),
        headers=HEADERS,
        params={filter_col: f"gte.{min_val}"},
    )
    r.raise_for_status()


def sb_insert(table: str, rows: list[dict]) -> None:
    """Insere lista de dicts em lotes de 200."""
    BATCH = 200
    for i in range(0, len(rows), BATCH):
        chunk = rows[i : i + BATCH]
        r = requests.post(_url(table), headers=HEADERS, json=chunk)
        if not r.ok:
            raise RuntimeError(f"{table} insert error {r.status_code}: {r.text[:300]}")


def sb_upsert(table: str, row: dict) -> None:
    """Upsert de uma única linha (usa Prefer: resolution=merge-duplicates)."""
    headers = {**HEADERS, "Prefer": "resolution=merge-duplicates,return=minimal"}
    r = requests.post(_url(table), headers=headers, json=[row])
    if not r.ok:
        raise RuntimeError(f"{table} upsert error {r.status_code}: {r.text[:300]}")


# ── Parse da planilha ─────────────────────────────────────────────────────────

def parse_triagem(ws) -> tuple[list[dict], list[str], str]:
    """Aba 'Triagem' → (demandas, alertas, enc_text)."""
    demandas = []
    alertas  = []
    enc_text = ""
    in_alerts = in_enc = False

    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        if all(c is None for c in row):
            continue
        first = str(row[0] or "").strip()

        if "ALERTAS" in first:
            in_alerts, in_enc = True, False
            continue
        if "ENCAMINHAMENTO" in first:
            in_alerts, in_enc = False, True
            continue

        if in_alerts:
            if first:
                alertas.append(first)
            continue
        if in_enc:
            if first:
                enc_text += first + " "
            continue

        if isinstance(row[0], (int, float)):
            demandas.append({
                "num"    : int(row[0]),
                "cliente": str(row[1] or ""),
                "dept"   : str(row[2] or ""),
                "cat"    : str(row[3] or ""),
                "sub"    : str(row[4] or ""),
                "func"   : str(row[5] or ""),
                "prazo"  : str(row[6] or ""),
                "prio"   : str(row[7] or ""),
                "dados"  : str(row[8] or ""),
                "pend"   : str(row[9] or ""),
                "dom"    : str(row[10] or "") if len(row) > 10 else "",
            })

    return demandas, alertas, enc_text.strip()


def parse_resumo(ws) -> dict:
    """Aba 'Resumo' → dict com total, by_dept, by_prio, data_hora."""
    out = {"total": 0, "by_dept": {}, "by_prio": {}, "data_hora": ""}
    mode = None

    for row in ws.iter_rows(values_only=True):
        if all(c is None for c in row):
            continue
        k = str(row[0] or "").strip()
        v = row[1] if len(row) > 1 else None

        if k.startswith("Data/hora"):
            out["data_hora"] = re.sub(r"^Data/hora geração:\s*", "", k, flags=re.I).strip() or str(v or "")
            continue
        if k == "Total de demandas:":
            out["total"] = int(v) if isinstance(v, (int, float)) else 0
            continue
        if k == "Por Departamento":
            mode = "dept"
            continue
        if k == "Por Prioridade":
            mode = "prio"
            continue
        if k == "Quantidade":
            continue

        if mode == "dept" and k and isinstance(v, (int, float)):
            out["by_dept"][k] = int(v)
        if mode == "prio" and k and isinstance(v, (int, float)):
            out["by_prio"][k] = int(v)

    return out


def _norm_name(s: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]", " ", s.lower())).strip()


def parse_triagem_completa(ws) -> dict:
    """Aba 'Triagem Completa' → dict keyed by contact_key."""
    convs: dict[str, dict] = {}
    cur_key = None

    MSG_RE = re.compile(
        r"^(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}:\d{2})\s*(?:>>>)?\s*\[(CLIENTE|ATENDENTE)\]\s*(.*)",
        re.I,
    )

    for row in ws.iter_rows(values_only=True):
        line = str(row[0] or "").strip()
        if not line:
            continue

        m_contact = re.match(r"CONTATO\s*:\s*(.+)", line, re.I)
        if m_contact:
            name_raw = re.split(r"\s*[|(+]", m_contact.group(1))[0].strip()
            cur_key  = _norm_name(name_raw)
            if cur_key not in convs:
                convs[cur_key] = {"raw_name": name_raw, "aba": "", "msgs": []}
            continue

        m_aba = re.match(r"ABA\s*:\s*(.+)", line, re.I)
        if m_aba and cur_key:
            convs[cur_key]["aba"] = m_aba.group(1).strip()
            continue

        if cur_key is None:
            continue
        if re.match(r"^[-=★]+$", line.replace(" ", "")):
            continue
        if re.match(
            r"^(PENDENTES|EM ATENDIMENTO|DIA:|Capturado em:|GESTTA|Última atualização|INSTRUÇÕES|Mantenha|={5,})",
            line, re.I,
        ):
            cur_key = None
            continue
        if line == "(nenhuma mensagem no período)":
            convs[cur_key]["msgs"].append({"type": "system", "time": "", "text": "(sem mensagens no período)"})
            continue

        m_msg = MSG_RE.match(line)
        if m_msg:
            date, time_, role, text = m_msg.groups()
            convs[cur_key]["msgs"].append({
                "type": role.lower(),
                "time": f"{date} {time_}",
                "text": text.strip(),
            })
        elif convs[cur_key]["msgs"]:
            last = convs[cur_key]["msgs"][-1]
            if last["type"] != "system":
                last["text"] += "\n" + line

    return convs


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    # Localizar planilha
    path = XLSX_PATH
    if not path.exists():
        # Procura recursiva no Desktop como fallback
        candidates = list(Path.home().glob("Desktop/**/triagem_atendimento.xlsx"))
        if not candidates:
            log.error("Arquivo não encontrado: %s", path)
            log.error("Coloque a planilha em: %s", XLSX_PATH)
            sys.exit(1)
        path = candidates[0]
        log.warning("Planilha encontrada em caminho alternativo: %s", path)

    log.info("Planilha: %s", path)

    # Importar openpyxl aqui para dar mensagem amigável se não instalado
    try:
        import openpyxl
    except ImportError:
        log.error("Instale a dependência:  pip install openpyxl requests")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # ── Parse ──────────────────────────────────────────────────────────────
    log.info("Lendo aba 'Triagem'…")
    ws_tri = wb["Triagem"] if "Triagem" in wb.sheetnames else None
    if ws_tri is None:
        log.error("Aba 'Triagem' não encontrada na planilha.")
        sys.exit(1)
    demandas, alertas, enc_text = parse_triagem(ws_tri)
    log.info("  %d demanda(s), %d alerta(s)", len(demandas), len(alertas))

    log.info("Lendo aba 'Resumo'…")
    ws_res = wb["Resumo"] if "Resumo" in wb.sheetnames else None
    resumo = parse_resumo(ws_res) if ws_res else {"total": len(demandas), "by_dept": {}, "by_prio": {}, "data_hora": ""}

    completa_name = next((n for n in wb.sheetnames if "completa" in n.lower()), None)
    log.info("Lendo aba '%s'…", completa_name or "Triagem Completa (não encontrada)")
    ws_comp = wb[completa_name] if completa_name else None
    convs   = parse_triagem_completa(ws_comp) if ws_comp else {}
    log.info("  %d conversa(s)", len(convs))

    wb.close()

    # ── Upload ─────────────────────────────────────────────────────────────
    now_iso = datetime.now(timezone.utc).isoformat()

    log.info("Limpando tabelas antigas…")
    try:
        sb_delete_all("triagem_demandas")
        sb_delete_all("triagem_resumo")
        sb_delete_all("triagem_conversas")
    except requests.HTTPError as e:
        log.error("Erro ao limpar tabelas: %s", e)
        sys.exit(1)

    log.info("Enviando %d demanda(s)…", len(demandas))
    if demandas:
        for d in demandas:
            d["uploaded_at"] = now_iso
        try:
            sb_insert("triagem_demandas", demandas)
        except RuntimeError as e:
            log.error(e)
            sys.exit(1)

    log.info("Enviando resumo…")
    resumo_row = {
        "id"         : 1,
        "total"      : resumo["total"] or len(demandas),
        "by_dept"    : resumo["by_dept"],
        "by_prio"    : resumo["by_prio"],
        "data_hora"  : resumo["data_hora"],
        "alerts"     : alertas,
        "enc_text"   : enc_text,
        "uploaded_at": now_iso,
    }
    try:
        sb_upsert("triagem_resumo", resumo_row)
    except RuntimeError as e:
        log.error(e)
        sys.exit(1)

    log.info("Enviando %d conversa(s)…", len(convs))
    if convs:
        conv_rows = [
            {
                "contact_key": key,
                "raw_name"   : c["raw_name"],
                "aba"        : c["aba"],
                "msgs"       : c["msgs"],
                "uploaded_at": now_iso,
            }
            for key, c in convs.items()
        ]
        try:
            sb_insert("triagem_conversas", conv_rows)
        except RuntimeError as e:
            log.error(e)
            sys.exit(1)

    log.info("✅ Upload concluído — %d demanda(s), %d conversa(s)", len(demandas), len(convs))


if __name__ == "__main__":
    main()
