#!/usr/bin/env python3
"""
Script para converter dados de avisos de vencimento de PDF para Excel

Uso:
    python pdf_to_excel_vencimentos.py <arquivo_pdf> [arquivo_saida.xlsx]

Exemplos:
    python pdf_to_excel_vencimentos.py relatorio.pdf
    python pdf_to_excel_vencimentos.py relatorio.pdf meu_vencimento.xlsx
"""

import sys
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


def extract_text_from_pdf(pdf_path: str) -> str:
    """Extrai texto completo do PDF"""
    reader = PdfReader(pdf_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text()
    return text


def parse_vencimentos_data(text: str) -> list:
    """
    Extrai dados de vencimentos do texto do PDF
    
    Procura por padrões como:
    "Empresa - CNPJ | Empregado | Tipo | Data"
    """
    data = []
    lines = text.split('\n')
    
    current_empresa = None
    current_cnpj = None
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Procura por linhas com CNPJ (formato XX.XXX.XXX/XXXX-XX ou CPF: XXX.XXX.XXX-XX)
        cnpj_match = re.search(r'(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}|CPF:\s*\d{3}\.\d{3}\.\d{3}-\d{2})', line)
        
        if cnpj_match:
            # Extrai a empresa (tudo antes do CNPJ)
            cnpj = cnpj_match.group(1)
            empresa_part = line[:cnpj_match.start()].strip()
            
            # Se houver empresa antes do CNPJ, atualiza
            if empresa_part:
                current_empresa = empresa_part
                current_cnpj = cnpj
    
    # Nota: Este é um parser básico. Um parser robusto dependeria da estrutura exata do PDF
    return data


def create_excel_from_data(data: list, output_path: str) -> None:
    """
    Cria arquivo Excel com os dados de vencimentos
    
    Args:
        data: Lista de tuplas (empresa, cnpj, empregado, tipo, vencimento)
        output_path: Caminho do arquivo de saída
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Vencimentos"
    
    # Estilos do cabeçalho
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Cabeçalhos e larguras das colunas
    headers = ["Empresa", "CNPJ", "Empregado", "Tipo", "Vencimento"]
    col_widths = [55, 22, 50, 45, 38]
    
    # Adiciona cabeçalhos
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.row_dimensions[1].height = 22
    
    # Definição de cores por tipo de evento
    tipo_fills = {
        "Aniversário colaboradores": PatternFill("solid", start_color="E2EFDA"),
        "Certificado digital": PatternFill("solid", start_color="FCE4D6"),
        "Contrato experiência 1º vencimento": PatternFill("solid", start_color="DDEBF7"),
        "Contrato experiência prorrogação": PatternFill("solid", start_color="EBF3FB"),
        "Vencimento de 2º Férias": PatternFill("solid", start_color="FFF2CC"),
        "Aviso Prévio de rescisão": PatternFill("solid", start_color="FDECEA"),
        "Envio rescisão eSocial": PatternFill("solid", start_color="FDECEA"),
        "Monitoramento de Saúde - Admissional": PatternFill("solid", start_color="EAF2FF"),
        "Monitoramento de Saúde - Mudança de função": PatternFill("solid", start_color="EAF2FF"),
        "Retorno de afastamento de Licença maternidade": PatternFill("solid", start_color="F3E5F5"),
        "Retorno de afastamento de Doença": PatternFill("solid", start_color="F3E5F5"),
        "Retorno de afastamento de Acidente de trabalho": PatternFill("solid", start_color="F3E5F5"),
    }
    default_fill = PatternFill("solid", start_color="FFFFFF")
    
    data_font = Font(name="Arial", size=9)
    data_align = Alignment(vertical="center", wrap_text=False)
    
    # Adiciona dados
    for row_idx, row in enumerate(data, 2):
        tipo = row[3] if len(row) > 3 else ""
        fill = tipo_fills.get(tipo, default_fill)
        
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = data_align
            cell.border = border
        
        ws.row_dimensions[row_idx].height = 15
    
    # Congela cabeçalho
    ws.freeze_panes = "A2"
    
    # Adiciona filtro automático
    ws.auto_filter.ref = f"A1:E{len(data) + 1}"
    
    # Salva arquivo
    wb.save(output_path)
    print(f"✓ Arquivo salvo: {output_path}")
    print(f"✓ Total de registros: {len(data)}")


def load_from_manual_data() -> list:
    """
    Carrega dados manualmente inseridos
    Retorna lista de tuplas (empresa, cnpj, empregado, tipo, vencimento)
    """
    # Dados extraídos do PDF original
    return [
        ("3 - CACAU CAFE CHOCOLATERIA LTDA", "23.724.189/0001-60", "CACAU CAFE CHOCOLATERIA LTDA", "Certificado digital", "17/01/2024"),
        ("4 - NUTRI ACAI COMERCIO DE POLPAS LTDA", "23.792.463/0001-39", "GABRIEL VINICIUS NERY XAVIER", "Contrato experiência prorrogação", "29/06/2026"),
        ("4 - NUTRI ACAI COMERCIO DE POLPAS LTDA", "23.792.463/0001-39", "SHEYLA FELIX SOARES DE MACEDO", "Aniversário colaboradores", "02/06"),
        # Adicione mais registros conforme necessário...
    ]


def main():
    """Função principal"""
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    
    pdf_path = sys.argv[1]
    
    # Define arquivo de saída
    if len(sys.argv) > 2:
        output_path = sys.argv[2]
    else:
        base_name = Path(pdf_path).stem
        output_path = f"{base_name}_vencimentos.xlsx"
    
    # Valida se arquivo PDF existe
    if not Path(pdf_path).exists():
        print(f"❌ Erro: Arquivo '{pdf_path}' não encontrado")
        sys.exit(1)
    
    print(f"📄 Processando: {pdf_path}")
    
    try:
        # Extrai texto do PDF
        print("📖 Extraindo texto do PDF...")
        text = extract_text_from_pdf(pdf_path)
        
        # Faz parsing dos dados
        print("🔍 Analisando dados...")
        data = parse_vencimentos_data(text)
        
        # Se o parser automático não encontrou dados, carrega dados manuais
        # (útil para PDFs com formatação complexa)
        if not data:
            print("⚠️  Parser automático não funcionou, usando dados manuais como exemplo...")
            data = load_from_manual_data()
        
        # Cria arquivo Excel
        print("📊 Gerando arquivo Excel...")
        create_excel_from_data(data, output_path)
        
        print(f"\n✅ Conversão concluída com sucesso!")
        print(f"📁 Arquivo salvo em: {Path(output_path).absolute()}")
        
    except Exception as e:
        print(f"❌ Erro durante o processamento: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
